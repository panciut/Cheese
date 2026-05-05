"""Build a unified image<->comment table for the Trentingrana dataset.

Output: data/unified_dataset.csv

One row per (image_path, comment) pair. Many-to-many: one image (one wheel
photographed in one session) typically maps to several panelist comments
across the 7 sensory attributes; if no comment matches, the image still
appears with empty comment fields (left join).
"""

from __future__ import annotations

import csv
import os
import re
import shutil
from datetime import datetime, date
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent / "data"
IMG_ROOT = ROOT / "TrentinGrana"
COMM_DIR = ROOT / "GT commenti liberi"
CODIF_DIR = COMM_DIR / "codifiche"
FLAT_DIR = ROOT / "images_flat"
OUT = ROOT / "intermediate" / "unified_dataset.csv"


def flat_name(rel_under_imgroot: Path) -> str:
    """Encode a path under TrentinGrana/ into a unique flat filename.

    Only the path separator is rewritten — spaces and other chars are kept
    so that filenames differing only by space-vs-underscore stay distinct.
    """
    return str(rel_under_imgroot).replace(os.sep, "__")

ATTRS_NORM = {
    "profumo": "Profumo",
    "sapore": "Sapore",
    "aroma": "Aroma",
    "texture": "Texture",
    "spessore della crosta": "Spessore della Crosta",
    "struttura della pasta": "Struttura della Pasta",
    "colore della pasta": "Colore della Pasta",
}


# ---------- codebook ----------
def load_dairy_map() -> dict[str, str]:
    """Returns dairy_norm (e.g. 'TN_302') -> product code (e.g. 'C0A')."""
    wb = openpyxl.load_workbook(CODIF_DIR / "codifica caseifici.xlsx", data_only=True)
    ws = wb["codici caseifici"]
    m: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        dairy, code = str(row[0]).strip(), str(row[1]).strip() if row[1] else None
        if not code:
            continue
        m[dairy] = code  # 'TN_302' -> 'C0A'
        # also accept compact form 'TN302' and bare '302'
        m[dairy.replace("_", "")] = code
        if dairy.startswith("TN_"):
            m[dairy[3:]] = code
    return m


def load_session_dates_2018() -> dict[int, date]:
    wb = openpyxl.load_workbook(COMM_DIR / "Commenti TOT_2018.xlsx", data_only=True)
    ws = wb["date_sedute_2018"]
    out: dict[int, date] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        n = int(row[0])
        d = row[1]
        if isinstance(d, datetime):
            out[n] = d.date()
        elif isinstance(d, date):
            out[n] = d
    return out


# ---------- comments ----------
def parse_score(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def load_comments() -> list[dict]:
    """Return long-format comments: one row per panelist*attribute*sample."""
    rows: list[dict] = []
    sess_dates_2018 = load_session_dates_2018()

    # 2018 file: cols Sogg, Seduta, Prod, <score>, Commenti
    f = COMM_DIR / "Commenti TOT_2018.xlsx"
    wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
    for sheet in wb.sheetnames:
        attr = ATTRS_NORM.get(sheet.strip().lower())
        if attr is None:
            continue
        ws = wb[sheet]
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                header = row
                continue
            if not row or row[0] is None:
                continue
            sogg, seduta, prod, score, comment = (row + (None,) * 5)[:5]
            try:
                seduta_n = int(seduta) if seduta is not None else None
            except (TypeError, ValueError):
                seduta_n = None
            d = sess_dates_2018.get(seduta_n) if seduta_n else None
            rows.append({
                "attribute": attr,
                "date": d.isoformat() if d else None,
                "session_num": seduta_n,
                "bimester": None,
                "production_date": None,
                "panelist": str(sogg).strip() if sogg else None,
                "product_code": str(prod).strip() if prod else None,
                "score": parse_score(score),
                "comment": str(comment).strip() if comment else None,
                "source_file": f.name,
            })

    # Later files: Data Seduta, N° Seduta, Bimestre, Data Produzione, Panelista, Prodotto, Commenti
    for fname in ["Commenti liberi_QTG_2019.xlsx",
                  "Commenti liberi_QTG_2020.xlsx",
                  "Commenti liberi_TEST_2021.xlsx"]:
        f = COMM_DIR / fname
        wb = openpyxl.load_workbook(f, data_only=True, read_only=True)
        for sheet in wb.sheetnames:
            attr = ATTRS_NORM.get(sheet.strip().lower())
            if attr is None:
                continue
            ws = wb[sheet]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue
                if not row or all(v is None for v in row):
                    continue
                data_sed, n_sed, bim, data_prod, panel, prod, comment = (row + (None,) * 7)[:7]
                if str(panel).strip() == "N/A":
                    continue
                d = None
                if isinstance(data_sed, datetime):
                    d = data_sed.date()
                elif isinstance(data_sed, date):
                    d = data_sed
                try:
                    sn = int(n_sed) if n_sed is not None else None
                except (TypeError, ValueError):
                    sn = None
                pd_str = None
                if isinstance(data_prod, (datetime, date)):
                    pd_str = data_prod.isoformat() if isinstance(data_prod, date) else data_prod.date().isoformat()
                elif data_prod is not None:
                    pd_str = str(data_prod).strip()
                c = str(comment).strip() if comment else None
                if c:
                    c = c.replace("\xa0", " ")
                rows.append({
                    "attribute": attr,
                    "date": d.isoformat() if d else None,
                    "session_num": sn,
                    "bimester": str(bim).strip() if bim else None,
                    "production_date": pd_str,
                    "panelist": str(panel).strip() if panel else None,
                    "product_code": str(prod).strip() if prod else None,
                    "score": None,
                    "comment": c,
                    "source_file": f.name,
                })
    return rows


# ---------- images ----------
DATE_PATTERNS = [
    re.compile(r"(?P<y>\d{4})[-_](?P<m>\d{2})[-_](?P<d>\d{2})"),  # 2018-08-29
    re.compile(r"(?P<d>\d{2})[-_](?P<m>\d{2})[-_](?P<y>\d{4})"),  # 04-09-2019 / 03_02_2021
]
ROMAN_RE = re.compile(r"\b(I|II|III|IV|V|VI|VII|VIII|IX|X)\b\s*bimestre", re.IGNORECASE)
SESSION_RE = re.compile(r"(\d+)\s*°?\s*Seduta", re.IGNORECASE)
DAIRY_TN_RE = re.compile(r"TN[_]?(\d{3})")
DAIRY_BARE_RE = re.compile(r"(?<![A-Za-z0-9])(\d{3})(?![A-Za-z0-9])")
SLOT_RE = re.compile(r"^P(\d+)([ab])", re.IGNORECASE)


def parse_image_path(path: Path) -> dict | None:
    """Extract metadata from an image path under TrentinGrana/."""
    rel = path.relative_to(IMG_ROOT)
    parts = list(rel.parts)
    if len(parts) < 2:
        return None
    year_folder = parts[0]  # e.g. '2018-2019_Trentingrana'

    # date: scan path components from deepest upward
    iso_date = None
    for comp in reversed(parts[:-1]):
        for pat in DATE_PATTERNS:
            m = pat.search(comp)
            if m:
                try:
                    iso_date = date(int(m["y"]), int(m["m"]), int(m["d"])).isoformat()
                except ValueError:
                    iso_date = None
                break
        if iso_date:
            break

    # bimester
    bimester = None
    for comp in parts[:-1]:
        m = ROMAN_RE.search(comp)
        if m:
            bimester = m.group(1).upper()
            break

    # session number
    session_num = None
    for comp in parts[:-1]:
        m = SESSION_RE.search(comp)
        if m:
            try:
                session_num = int(m.group(1))
            except ValueError:
                pass
            break

    fname = path.name
    slot = SLOT_RE.search(fname)
    panel_slot = (int(slot.group(1)), slot.group(2).lower()) if slot else (None, None)

    dairy = None
    md = DAIRY_TN_RE.search(fname)
    if md:
        dairy = "TN_" + md.group(1)
    else:
        md = DAIRY_BARE_RE.search(fname)
        if md:
            dairy = "TN_" + md.group(1)

    view = None
    fl = fname.lower()
    if "fetta" in fl:
        view = "Fetta"
    elif "grana" in fl:
        view = "Grana"

    return {
        "image_path": str(path),
        "image_filename": fname,
        "year_folder": year_folder,
        "session_date": iso_date,
        "session_num_img": session_num,
        "bimester_img": bimester,
        "panel_slot": panel_slot[0],
        "panel_replicate": panel_slot[1],
        "dairy_id": dairy,
        "view": view,
    }


def walk_images() -> list[dict]:
    out: list[dict] = []
    for p in IMG_ROOT.rglob("*.bmp"):
        if p.name.lower() in {"thumbs.db"}:
            continue
        meta = parse_image_path(p)
        if meta is None:
            continue
        out.append(meta)
    return out


# ---------- join ----------
def main():
    dairy_map = load_dairy_map()
    images = walk_images()
    # attach product_code
    for im in images:
        im["product_code"] = dairy_map.get(im["dairy_id"]) if im["dairy_id"] else None

    comments = load_comments()
    # index comments by (date, product_code)
    by_dp: dict[tuple[str | None, str | None], list[dict]] = {}
    for c in comments:
        key = (c["date"], c["product_code"])
        by_dp.setdefault(key, []).append(c)

    cols = [
        "image_path", "image_path_flat", "image_filename", "view",
        "year_folder", "session_date", "session_num", "bimester",
        "panel_slot", "panel_replicate", "dairy_id", "product_code",
        "panelist", "attribute", "score", "comment",
        "production_date", "comment_source_file",
    ]
    FLAT_DIR.mkdir(parents=True, exist_ok=True)
    n_imgs = 0
    n_paired = 0
    n_orphan_imgs = 0
    n_rows = 0
    n_copied = 0
    n_skipped_existing = 0
    seen_flat: set[str] = set()
    with OUT.open("w", newline="") as fh:
        w = csv.DictWriter(fh, fieldnames=cols)
        w.writeheader()
        for im in images:
            n_imgs += 1
            key = (im["session_date"], im["product_code"])
            matches = by_dp.get(key, [])
            src = Path(im["image_path"])
            rel_path = str(src.relative_to(ROOT.parent))
            flat = flat_name(src.relative_to(IMG_ROOT))
            if flat in seen_flat:
                raise RuntimeError(f"flat name collision: {flat}")
            seen_flat.add(flat)
            dst = FLAT_DIR / flat
            if dst.exists() and dst.stat().st_size == src.stat().st_size:
                n_skipped_existing += 1
            else:
                shutil.copy2(src, dst)
                n_copied += 1
            flat_rel = str(dst.relative_to(ROOT.parent))
            if not matches:
                n_orphan_imgs += 1
                w.writerow({
                    "image_path": rel_path,
                    "image_path_flat": flat_rel,
                    "image_filename": im["image_filename"],
                    "view": im["view"],
                    "year_folder": im["year_folder"],
                    "session_date": im["session_date"],
                    "session_num": im["session_num_img"],
                    "bimester": im["bimester_img"],
                    "panel_slot": im["panel_slot"],
                    "panel_replicate": im["panel_replicate"],
                    "dairy_id": im["dairy_id"],
                    "product_code": im["product_code"],
                    "panelist": None, "attribute": None, "score": None,
                    "comment": None, "production_date": None,
                    "comment_source_file": None,
                })
                n_rows += 1
                continue
            n_paired += 1
            for c in matches:
                w.writerow({
                    "image_path": rel_path,
                    "image_path_flat": flat_rel,
                    "image_filename": im["image_filename"],
                    "view": im["view"],
                    "year_folder": im["year_folder"],
                    "session_date": im["session_date"],
                    "session_num": im["session_num_img"] or c["session_num"],
                    "bimester": im["bimester_img"] or c["bimester"],
                    "panel_slot": im["panel_slot"],
                    "panel_replicate": im["panel_replicate"],
                    "dairy_id": im["dairy_id"],
                    "product_code": im["product_code"],
                    "panelist": c["panelist"],
                    "attribute": c["attribute"],
                    "score": c["score"],
                    "comment": c["comment"],
                    "production_date": c["production_date"],
                    "comment_source_file": c["source_file"],
                })
                n_rows += 1

    # diagnostics
    by_year_paired: dict[str, list[int]] = {}
    for im in images:
        key = (im["session_date"], im["product_code"])
        paired = 1 if by_dp.get(key) else 0
        by_year_paired.setdefault(im["year_folder"], [0, 0])
        by_year_paired[im["year_folder"]][0] += 1
        by_year_paired[im["year_folder"]][1] += paired

    print(f"images walked       : {n_imgs}")
    print(f"images with comments: {n_paired}")
    print(f"orphan images       : {n_orphan_imgs}")
    print(f"output rows         : {n_rows}")
    print(f"comments loaded     : {len(comments)}")
    print()
    print("per-year-folder pairing rate:")
    for y, (tot, ok) in sorted(by_year_paired.items()):
        pct = 100 * ok / tot if tot else 0
        print(f"  {y}: {ok}/{tot}  ({pct:.1f}%)")

    # diagnose unmapped dairies
    unmapped = sorted({im["dairy_id"] for im in images if im["dairy_id"] and not im["product_code"]})
    if unmapped:
        print(f"\nimage dairy IDs not in codebook: {unmapped[:20]}{'...' if len(unmapped) > 20 else ''}")
    no_dairy = sum(1 for im in images if not im["dairy_id"])
    no_date = sum(1 for im in images if not im["session_date"])
    print(f"\nimages w/o dairy parsed  : {no_dairy}")
    print(f"images w/o date parsed   : {no_date}")
    print(f"\nflat copies: copied={n_copied}, already-present={n_skipped_existing}, dir={FLAT_DIR}")
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
